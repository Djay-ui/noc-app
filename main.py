import csv
import hashlib
import hmac
import io
import json
import os
import smtplib
from datetime import datetime, timedelta, timezone
from email import encoders
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Optional

import psycopg2
from fastapi import (BackgroundTasks, Depends, FastAPI, File, Form,
                     HTTPException, Query, Request, UploadFile)
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from psycopg2.extras import RealDictCursor
from pydantic import BaseModel, Field

from email_helpers import build_email_message

app = FastAPI(title="Teleglobal NOC Automation Platform")
templates = Jinja2Templates(directory="templates")

# Configuration Constants
SMTP_SERVER = "mail.teleglobal.in"
SMTP_PORT = 465
SMTP_USER = "noc@teleglobal.in"
SMTP_PASSWORD = "8QKti-lme88&"
GLOBAL_MANDATORY_CC = ["noc@teleglobal.in"]

AUTH_SECRET_KEY = "SUPER_SECRET_NOC_KEY_2026_TRACK_SYSTEM_SECURE"
COOKIE_NAME = "noc_session_token"
PASSWORD_SALT = "noc_salt_2026"

def get_db_connection():
    return psycopg2.connect(
        dbname="noc_ticketing", 
        user="noc_admin", 
        password="SecureNocPassword2026!", 
        host="localhost"
    )

# Async Background Mail Dispatcher (Fixes App Hanging)
def send_smtp_email_background(msg_string: str, all_recipients: list):
    try:
        server = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, timeout=10)
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(SMTP_USER, all_recipients, msg_string)
        server.quit()
    except Exception as smtp_err:
        print(f"Background SMTP Dispatch Failure: {str(smtp_err)}")

# SCALABLE ALARM CONFIGURATION MATRIX

from datetime import datetime, timedelta, timezone

from fastapi import (BackgroundTasks, Depends, FastAPI, File, Form,
                     HTTPException, UploadFile)

# --- SCALABLE ALARM CONFIGURATION MATRIX ---

def determine_email_template(issue_category: str, status: str) -> str:
    if not issue_category:
        cat = ""
    else:
        cat = issue_category.strip().lower()

    # Exact matching logic (case-insensitive & whitespace trimmed)
    if cat in ["primary default update", "primary default", "primary", "general issue"]:
        return "primary_up.html" if status == "Closed" else "primary.html"
    elif cat == "ip is not pingable":
        return "ip_up.html" if status == "Closed" else "ip_down.html"
    elif cat in ["switch isolated", "switch is isolated"]:
        return "switch_up.html" if status == "Closed" else "switch_down.html"
    elif cat in ["media device down", "media device"]:
        return "mkt_up.html" if status == "Closed" else "mkt.html"
    elif cat == "high bandwidth utilization":
        return "high.html"

    # Default fallbacks
    if status == "Closed":
        return "link_up.html"
    elif status == "In Monitoring":
        return "link_monitoring.html"
    else:
        return "link_down.html"

def assign_priority_and_sla(issue_category: str):
    now = datetime.utcnow()
    
    if issue_category in ["Core Backbone Down", "Complete Site Isolation", "Total Blackout"]:
        return "P1", now + timedelta(hours=1)
    elif issue_category in ["high bandwidth utilization", "High Packet Loss", "Latency Issues", "Redundancy Lost"]:
        return "P2", now + timedelta(hours=4)
    elif issue_category in ["Interface Flapping", "Minor Config Error"]:
        return "P3", now + timedelta(hours=24)
    else:
        return "P4", now + timedelta(days=3)

# Security & Crypto Helpers
def hash_password(password: str) -> str:
    pwd_bytes = password.encode('utf-8')
    salt_bytes = PASSWORD_SALT.encode('utf-8')
    dk = hashlib.pbkdf2_hmac('sha256', pwd_bytes, salt_bytes, 100000)
    return dk.hex()

def generate_session_token(user_row: dict) -> str:
    expires = (datetime.now(timezone.utc) + timedelta(hours=12)).isoformat()
    payload = {
        "id": user_row["id"],
        "username": user_row["username"],
        "full_name": user_row["full_name"],
        "role": user_row["role"],
        "expires": expires
    }
    payload_str = json.dumps(payload)
    signature = hmac.new(AUTH_SECRET_KEY.encode('utf-8'), payload_str.encode('utf-8'), hashlib.sha256).hexdigest()
    return f"{payload_str}.{signature}"

def verify_session_token(token: str) -> dict:
    if not token:
        return None
    try:
        payload_str, signature = token.rsplit('.', 1)
        expected_sig = hmac.new(AUTH_SECRET_KEY.encode('utf-8'), payload_str.encode('utf-8'), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected_sig):
            return None
        payload = json.loads(payload_str)
        if datetime.fromisoformat(payload["expires"]) < datetime.now(timezone.utc):
            return None
        return payload
    except Exception:
        return None

#signiture

import re


# --- SIGNATURE AUTOMATION HELPERS ---
def get_dynamic_signature(user_info: dict) -> str:
    """Generates a clean dynamic signature string mapping user parameters."""
    engineer_name = user_info.get("full_name", user_info.get("username", "NOC Specialist"))
    if "|" in engineer_name:
        engineer_name = engineer_name.split("|")[0].strip()
        
    designation = user_info.get("role", "NOC Engineer").capitalize()
    if designation.lower() == "admin":
        designation = "NOC Administrator"
        
    return f"""
    <br><br>
    <p style="font-family: Arial, sans-serif; font-size: 13px; color: #333333; line-height: 1.5;">
        Best Regards,<br>
        <strong style="color: #0b4d91;">{engineer_name}</strong><br>
        <span>{designation}</span><br>
        <span style="color: #555555;">Teleglobal Communications Pvt. Ltd.</span>
    </p>
    """

def append_signature(html_body: str, user_info: dict) -> str:
    """
    Injects operator details into the template placeholders and prevents 
    external file auto-appends. Only uses the signature built into the template.
    """
    engineer_name = user_info.get("full_name", user_info.get("username", "NOC Specialist"))
    if "|" in engineer_name:
        engineer_name = engineer_name.split("|")[0].strip()
        
    designation = user_info.get("role", "NOC Engineer").capitalize()
    if designation.lower() == "admin":
        designation = "NOC Administrator"

    # Replace both lowercase and uppercase variations safely
    html_body = html_body.replace("{operator_name}", engineer_name)
    html_body = html_body.replace("{OPERATOR_NAME}", engineer_name)
    html_body = html_body.replace("{designation}", designation)
    html_body = html_body.replace("{DESIGNATION}", designation)
    
    return html_body

# Dependency Providers
async def get_current_user(request: Request):
    token = request.cookies.get(COOKIE_NAME)
    user = verify_session_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized session.")
    return user

async def get_optional_user(request: Request):
    token = request.cookies.get(COOKIE_NAME)
    return verify_session_token(token)

class CircuitModel(BaseModel):
    circuit_id: str
    customer_name: str
    company_name: str
    customer_email: str
    phone_number: str
    address: str

class FiberModel(BaseModel):
    id: Optional[int] = None
    fiber_name: str
    fiber_vendor: str
    contact_numbers: Optional[str] = None
    email_address: Optional[str] = None
    route_details: Optional[str] = None

class UserUpdateModel(BaseModel):
    user_id: int
    username: str
    email_id: str
    full_name: str
    employee_id: str
    role: str
    password: str = None

class UserCreateModel(BaseModel):
    username: str
    email_id: str
    full_name: str
    employee_id: str
    role: str
    password: str

class ChangePasswordPayload(BaseModel):
    current_password: str
    new_password: str

class ReportPayload(BaseModel):
    report_type: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    team: Optional[str] = None

###############################################################################################

# Base payload definitions used during core requests
class TicketBase(BaseModel):
    circuit_id: str
    issue_category: str
    root_cause_segment: Optional[str] = None
    status: str = "Open"
    assigned_team: str
    open_by_name: str
    priority: str = Field(default="P3", description="P1 (Critical) through P4 (Informational) SLA levels")
    sla_deadline: Optional[datetime] = None

# Input Validation Model mapping for HTTP POST routes
class TicketCreate(TicketBase):
    pass

# Response Marshalling Schema mapping for analytical JSON responses
class TicketResponse(TicketBase):
    ticket_id: int
    closed_by_name: Optional[str] = None
    created_at: datetime
    closed_at: Optional[datetime] = None
    resolution_minutes: Optional[int] = None
    is_sla_breached: bool = False

    class Config:
        from_attributes = True  # Handles native ORM row object mappings

# Page Routing Interceptors
@app.get("/login", response_class=HTMLResponse)
async def route_login_page(request: Request, user=Depends(get_optional_user)):
    if user:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(request=request, name="login.html")

@app.get("/", response_class=HTMLResponse)
async def route_dashboard(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="dashboard.html", context={"user": user})

# Place your roster route right here alongside other UI page routes:

from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse

# App base directory (/opt/noc-app)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Save the JSON file in /opt/noc-app/roster_data.json
ROSTER_STORAGE_FILE = os.path.join(BASE_DIR, "roster_data.json")

def load_roster_data():
    if os.path.exists(ROSTER_STORAGE_FILE):
        try:
            with open(ROSTER_STORAGE_FILE, "r") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading roster data: {e}")
            return {}
    return {}

def save_roster_data_to_disk(data):
    try:
        with open(ROSTER_STORAGE_FILE, "w") as f:
            json.dump(data, f, indent=4)
    except Exception as e:
        print(f"Error saving roster data: {e}")

# Data payload model for API
class RosterSavePayload(BaseModel):
    yearMonth: str
    rosterData: dict

# 1. Roster Page Route
@app.get("/roster", response_class=HTMLResponse)
async def route_roster_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="roster.html",
        context={"user": user}
    )

# 2. Get Roster Data API
@app.get("/api/roster")
async def get_roster(month: str, user=Depends(get_optional_user)):
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    all_data = load_roster_data()
    month_data = all_data.get(month, {})
    return JSONResponse(content=month_data)

# 3. Save Roster Data API

@app.post("/api/roster/save")
async def save_roster(payload: RosterSavePayload, user=Depends(get_optional_user)):
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Safely extract user role & username whether dict or object
    user_role = user.get('role', '') if isinstance(user, dict) else getattr(user, 'role', '')
    username = user.get('username', '') if isinstance(user, dict) else getattr(user, 'username', '')
    
    if user_role != 'admin' and username != 'admin':
        raise HTTPException(status_code=403, detail="Only Admin can save changes.")

    all_data = load_roster_data()
    all_data[payload.yearMonth] = payload.rosterData
    save_roster_data_to_disk(all_data)

    return JSONResponse(content={"status": "success", "message": "Roster saved successfully!"})

@app.get("/tickets", response_class=HTMLResponse)
async def route_tickets_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="tickets.html", context={"user": user})

@app.get("/circuits", response_class=HTMLResponse)
async def route_circuits_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user["role"] != "admin":
        return HTMLResponse("<html><body><h3>Access Denied: Clearance Required</h3><a href='/'>Return Dashboard</a></body></html>", status_code=403)
    return templates.TemplateResponse(request=request, name="circuits.html", context={"user": user})

@app.get("/fiber", response_class=HTMLResponse)
async def route_fiber_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user["role"] != "admin":
        return HTMLResponse("<html><body><h3>Access Denied: Clearance Required</h3><a href='/'>Return Dashboard</a></body></html>", status_code=403)
    return templates.TemplateResponse(request=request, name="fiber.html", context={"user": user})

@app.get("/system-mail/welcome", response_class=HTMLResponse)
async def route_system_mail_welcome_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="system_mail_welcome.html", context={"user": user})

# Route to render the Access Mail UI page
@app.get("/system-mail/access", response_class=HTMLResponse)
async def route_system_mail_access_page(request: Request, user = Depends(get_optional_user)):
    """
    Renders the secure Access Request operational template dashboard interface.
    """
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="system_mail_access.html", context={"user": user})

@app.get("/system-mail/bandwidth", response_class=HTMLResponse)
async def route_bandwidth_change_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="system_mail_bandwidth.html", context={"user": user})

@app.get("/system-mail/terminate", response_class=HTMLResponse)
async def route_link_termination_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="system_mail_terminate.html", context={"user": user})

@app.get("/system-mail/rfo", response_class=HTMLResponse)
async def route_system_mail_rfo_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(
        request=request, 
        name="system_mail_rfo.html", 
        context={"user": user}
    )

# =========================================================================
# ROUTE TO RENDER CUSTOM MAIL PAGE
# =========================================================================
@app.get("/system-mail/custom", response_class=HTMLResponse)
async def route_system_mail_custom_page(request: Request, user=Depends(get_optional_user)):
    """Renders the Custom Mail dynamic operational interface."""
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="system_mail_custom.html", context={"user": user})

@app.get("/admin/users", response_class=HTMLResponse)
async def route_user_management_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user["role"] != "admin":
        return HTMLResponse("<html><body><h3>Access Denied</h3></body></html>", status_code=403)
    return templates.TemplateResponse(request=request, name="users.html", context={"user": user})

@app.get("/reports", response_class=HTMLResponse)
async def get_reports_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="reports.html", context={"user": user})

# Auth Actions
@app.post("/api/auth/login")
async def api_login(username: str = Form(...), password: str = Form(...)):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(%s)", (username.strip(),))
    user = cursor.fetchone()
    cursor.close(); conn.close()

    if not user or user["password_hash"] != hash_password(password):
        return HTMLResponse("<html><body><script>alert('Invalid operational credentials.'); window.location.href='/login';</script></body></html>", status_code=400)

    token = generate_session_token(user)
    
    # Redirect to password change page if it's their first login, otherwise go home
    redirect_target = "/change-password" if user.get("is_first_login", True) else "/"
    
    response = RedirectResponse(url=redirect_target, status_code=303)
    response.set_cookie(key=COOKIE_NAME, value=token, httponly=True, max_age=43200, samesite="lax")
    return response

@app.get("/api/auth/logout")
async def api_logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(COOKIE_NAME)
    return response

@app.get("/api/auth/me")
async def api_get_me(user=Depends(get_current_user)):
    return user

# =========================================================================
# WELCOME MAIL PIPELINE 
# =========================================================================
@app.post("/api/tools/send-welcome-mail")
async def api_send_provisioning_welcome_mail(
    background_tasks: BackgroundTasks,
    circuit_id: str = Form(...),
    company_name: str = Form(...),
    bandwidth_speed: str = Form(...),
    commissioning_date: str = Form(...),
    wan_ip_details: str = Form(...),
    usable_ips: str = Form(...),
    default_gateway: str = Form(...),
    subnet_mask: str = Form(...),
    customer_email: str = Form(...),
    cc_emails: str = Form(""),
    testing_snap: UploadFile = File(...),
    escalation_matrix: UploadFile = File(...), 
    user=Depends(get_current_user)
):
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))
    if "|" in engineer_identity:
        engineer_identity = engineer_identity.split("|")[0].strip()

    try:
        parsed_date = datetime.strptime(commissioning_date.strip(), "%Y-%m-%d")
        formatted_date = parsed_date.strftime("%d %B %Y")
    except Exception:
        formatted_date = commissioning_date.strip()

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        db_query = """
            INSERT INTO customers (circuit_id, customer_name, customer_email)
            VALUES (%s, %s, %s)
            ON CONFLICT (circuit_id) DO UPDATE SET
                customer_name = EXCLUDED.customer_name,
                customer_email = EXCLUDED.customer_email;
        """
        cursor.execute(db_query, (circuit_id.strip(), company_name.strip(), customer_email.strip()))
        conn.commit()
    except Exception as db_err:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database synchronization failed: {str(db_err)}")
    finally:
        cursor.close(); conn.close()

    try:
        template = templates.get_template("emails/welcome_mail.html")
        hydrated_body = template.render({
            "circuit_id": circuit_id.strip(),
            "company_name": company_name.strip(),
            "bandwidth_speed": bandwidth_speed.strip(),
            "commissioning_date": formatted_date,
            "wan_ip_details": wan_ip_details.strip(),
            "usable_ips": usable_ips.strip(),
            "default_gateway": default_gateway.strip(),
            "subnet_mask": subnet_mask.strip(),
            "operator_name": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(status_code=500, detail=f"Failed to compile template placeholders: {str(render_err)}")

    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = f"Welcome to TeleGlobal Communications Pvt. Ltd || {company_name.strip()} || {circuit_id.strip()}"
    
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip():
                cc_list.append(addr.strip())
    if cc_list:
        msg['Cc'] = ", ".join(cc_list)
        recipients.extend(cc_list)

    msg.attach(MIMEText(append_signature(hydrated_body, user), 'html'))

    if not escalation_matrix.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="The Escalation Matrix attachment must strictly be a PDF document.")
        
    try:
        matrix_bytes = await escalation_matrix.read()
        matrix_part = MIMEApplication(matrix_bytes)
        matrix_part['Content-Disposition'] = f'attachment; filename="{escalation_matrix.filename}"'
        msg.attach(matrix_part)
    except Exception as esc_err:
        raise HTTPException(status_code=500, detail=f"Failed compiling escalation matrix attachment: {str(esc_err)}")

    try:
        file_bytes = await testing_snap.read()
        snap_part = MIMEApplication(file_bytes)
        snap_part['Content-Disposition'] = f'attachment; filename="{testing_snap.filename}"'
        msg.attach(snap_part)
    except Exception as img_err:
         raise HTTPException(status_code=500, detail=f"Failed compiling network snapshot logs: {str(img_err)}")

    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)
    return {"status": "success", "message": "Welcome onboarding welcome pack sent successfully."}

# =========================================================================
# BANDWIDTH ALTERATION MANAGEMENT PIPELINE 
# =========================================================================
@app.post("/api/tools/send-bandwidth-change-mail")
async def api_send_bandwidth_change_mail(
    background_tasks: BackgroundTasks,
    change_type: str = Form(...),             
    circuit_id: str = Form(...),
    company_name: str = Form(...),
    old_bandwidth_speed: str = Form(...),
    new_bandwidth_speed: str = Form(...),
    effective_date: str = Form(...),          
    customer_email: str = Form(...),          
    cc_emails: str = Form(""),
    user=Depends(get_current_user)
):
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))
    if "|" in engineer_identity:
        engineer_identity = engineer_identity.split("|")[0].strip()

    try:
        parsed_date = datetime.strptime(effective_date.strip(), "%Y-%m-%d")
        formatted_date = parsed_date.strftime("%d %B %Y")
    except Exception:
        formatted_date = effective_date.strip()

    action_title = "Upgradation" if change_type.lower() == "upgrade" else "Downgradation"
    action_verb = "upgraded" if change_type.lower() == "upgrade" else "downgraded"

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        db_query = """
            INSERT INTO customers (circuit_id, customer_name, customer_email)
            VALUES (%s, %s, %s)
            ON CONFLICT (circuit_id) DO UPDATE SET
                customer_name = EXCLUDED.customer_name,
                customer_email = EXCLUDED.customer_email;
        """
        cursor.execute(db_query, (circuit_id.strip(), company_name.strip(), customer_email.strip()))
        conn.commit()
    except Exception as db_err:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database synchronization failed: {str(db_err)}")
    finally:
        cursor.close(); conn.close()

    try:
        template = templates.get_template("emails/bandwidth_change_mail.html")
        hydrated_body = template.render({
            "action_title": action_title,
            "action_verb": action_verb,
            "circuit_id": circuit_id.strip(),
            "company_name": company_name.strip(),
            "old_bandwidth_speed": old_bandwidth_speed.strip(),
            "new_bandwidth_speed": new_bandwidth_speed.strip(),
            "effective_date": formatted_date,
            "operator_name": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(status_code=500, detail=f"Failed to compile template placeholders: {str(render_err)}")

    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = f"Link Bandwidth {action_title} Confirmation || {company_name.strip()} || {circuit_id.strip()}"
    
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip():
                cc_list.append(addr.strip())
    if cc_list:
        msg['Cc'] = ", ".join(cc_list)
        recipients.extend(cc_list)

    msg.attach(MIMEText(append_signature(hydrated_body, user), 'html'))
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)
    log_operational_event("bandwidth_upgrade_logs", circuit_id.strip(), old_bandwidth_speed.strip(), new_bandwidth_speed.strip())
    return {"status": "success", "message": f"Bandwidth {action_title} notifications queued cleanly."}

# =========================================================================
# LINK TERMINATION MANAGEMENT PIPELINE
# =========================================================================
@app.post("/api/tools/send-termination-mail")
async def api_send_termination_mail(
    background_tasks: BackgroundTasks,
    circuit_id: str = Form(...),
    company_name: str = Form(...),
    termination_date: str = Form(...),        
    customer_email: str = Form(...),          
    cc_emails: str = Form(""),
    user=Depends(get_current_user)
):
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))
    if "|" in engineer_identity:
        engineer_identity = engineer_identity.split("|")[0].strip()

    try:
        parsed_date = datetime.strptime(termination_date.strip(), "%Y-%m-%d")
        formatted_date = parsed_date.strftime("%d %B %Y")
    except Exception:
        formatted_date = termination_date.strip()

    try:
        template = templates.get_template("emails/terminate.html")
        hydrated_body = template.render({
            "CIRCUIT_ID": circuit_id.strip(),
            "CUSTOMER_NAME": company_name.strip(),
            "TERMINATION_DATE": formatted_date,
            "ENGINEER_NAME": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(status_code=500, detail=f"Failed to find or parse emails/terminate.html: {str(render_err)}")

    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = f"Link Termination Notification || {company_name.strip()} || {circuit_id.strip()}"
    
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip():
                cc_list.append(addr.strip())
    if cc_list:
        msg['Cc'] = ", ".join(cc_list)
        recipients.extend(cc_list)

    msg.attach(MIMEText(append_signature(hydrated_body, user), 'html'))
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)
    log_operational_event("link_termination_logs", circuit_id.strip(), "Termination requested by client", None)
    return {"status": "success", "message": "Link termination announcement queued cleanly via background layers."}

# =========================================================================
# CORE DATA MANAGEMENT LAYER & API METRICS ENGINE
# =========================================================================
@app.get("/api/circuit/{circuit_id}")
async def get_circuit_details(circuit_id: str, user=Depends(get_current_user)):
    search_term = circuit_id.strip()
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    query = """
        SELECT circuit_id, customer_name, company_name, customer_email, phone_number, address 
        FROM customers 
        WHERE LOWER(circuit_id) LIKE LOWER(%s)
           OR LOWER(customer_name) LIKE LOWER(%s)
           OR LOWER(company_name) LIKE LOWER(%s)
           OR LOWER(customer_email) LIKE LOWER(%s)
           OR phone_number LIKE %s
           OR LOWER(address) LIKE LOWER(%s)
        ORDER BY (LOWER(circuit_id) = LOWER(%s)) DESC
    """
    wildcard_term = f"%{search_term}%"
    cursor.execute(query, (
        wildcard_term, wildcard_term, wildcard_term, 
        wildcard_term, wildcard_term, wildcard_term, search_term
    ))
    circuit_records = cursor.fetchall()
    cursor.close(); conn.close()
    
    if not circuit_records:
        raise HTTPException(status_code=404, detail="No matching customer profile or circuit was found.")
    return circuit_records

@app.get("/api/circuits/all")
async def api_get_all_circuits(search: str = "", user=Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if search:
            term = f"%{search.strip()}%"
            cursor.execute("""
                SELECT circuit_id, customer_name, company_name, customer_email, phone_number, address 
                FROM customers 
                WHERE LOWER(circuit_id) LIKE LOWER(%s)
                   OR LOWER(customer_name) LIKE LOWER(%s)
                   OR LOWER(company_name) LIKE LOWER(%s)
                   OR phone_number LIKE %s
                ORDER BY circuit_id ASC
            """, (term, term, term, term))
        else:
            cursor.execute("SELECT circuit_id, customer_name, company_name, customer_email, phone_number, address FROM customers ORDER BY circuit_id ASC")
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database fetch failure: {str(e)}")
    finally:
        cursor.close(); conn.close()

@app.post("/api/circuit/save")
@app.post("/api/circuits/save")
@app.put("/api/circuit/save")
@app.put("/api/circuits/save")
async def api_save_circuit(request: Request, user=Depends(get_current_user)):
    # 1. Authorization Guard
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin authorization required.")

    # 2. Flexible Payload Parsing (Handles JSON or Form Submissions)
    try:
        data = await request.json()
    except Exception:
        try:
            form_data = await request.form()
            data = dict(form_data)
        except Exception:
            raise HTTPException(
                status_code=400, 
                detail="Invalid request format. Payload must be valid JSON or Form data."
            )

    # 3. Extract Values from Payload
    circuit_id = data.get("circuit_id")
    customer_name = data.get("customer_name")
    company_name = data.get("company_name")
    customer_email = data.get("customer_email")
    phone_number = data.get("phone_number")
    address = data.get("address")

    # Validation Guard
    if not circuit_id:
        raise HTTPException(status_code=400, detail="Missing required field: circuit_id")

    # 4. Database Transaction
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Database Upsert query to elegantly handle updates vs inserts without conflicts
        query = """
            INSERT INTO customers (circuit_id, customer_name, company_name, customer_email, phone_number, address)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (circuit_id) 
            DO UPDATE SET
                customer_name = EXCLUDED.customer_name,
                company_name = EXCLUDED.company_name,
                customer_email = EXCLUDED.customer_email,
                phone_number = EXCLUDED.phone_number,
                address = EXCLUDED.address;
        """
        
        cursor.execute(query, (
            str(circuit_id).strip(),
            str(customer_name).strip() if customer_name else None,
            str(company_name).strip() if company_name else None,
            str(customer_email).strip() if customer_email else None,
            str(phone_number).strip() if phone_number else None,
            str(address).strip() if address else None
        ))
        
        conn.commit()
        return {"status": "success", "message": f"Customer record for circuit {circuit_id} has been saved successfully."}

    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database save execution failure: {str(e)}")
        
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.get("/api/fiber/all")
async def api_get_all_fiber(search: str = "", user=Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if search:
            term = f"%{search.strip()}%"
            cursor.execute("""
                SELECT id, fiber_name, fiber_vendor, contact_numbers, email_address, route_details 
                FROM fiber_db 
                WHERE LOWER(fiber_name) LIKE LOWER(%s)
                   OR LOWER(fiber_vendor) LIKE LOWER(%s)
                   OR contact_numbers LIKE %s
                   OR LOWER(email_address) LIKE LOWER(%s)
                ORDER BY fiber_name ASC
            """, (term, term, term, term))
        else:
            cursor.execute("SELECT id, fiber_name, fiber_vendor, contact_numbers, email_address, route_details FROM fiber_db ORDER BY fiber_name ASC")
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database fetch failure: {str(e)}")
    finally:
        cursor.close()
        conn.close()

@app.post("/api/fiber/save")
@app.put("/api/fiber/save")
async def api_save_fiber(request: Request, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin authorization required.")

    try:
        data = await request.json()
    except Exception:
        try:
            form_data = await request.form()
            data = dict(form_data)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid request format.")

    fiber_id = data.get("id")
    fiber_name = data.get("fiber_name")
    fiber_vendor = data.get("fiber_vendor")
    contact_numbers = data.get("contact_numbers")
    email_address = data.get("email_address")
    route_details = data.get("route_details")

    if not fiber_name or not fiber_vendor:
        raise HTTPException(status_code=400, detail="Missing required fields: fiber_name and fiber_vendor are mandatory.")

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if fiber_id:
            cursor.execute("""
                UPDATE fiber_db
                SET fiber_name = %s, fiber_vendor = %s, contact_numbers = %s, email_address = %s, route_details = %s
                WHERE id = %s
            """, (str(fiber_name).strip(), str(fiber_vendor).strip(), str(contact_numbers).strip() if contact_numbers else None, str(email_address).strip() if email_address else None, str(route_details).strip() if route_details else None, fiber_id))
        else:
            cursor.execute("""
                INSERT INTO fiber_db (fiber_name, fiber_vendor, contact_numbers, email_address, route_details)
                VALUES (%s, %s, %s, %s, %s)
            """, (str(fiber_name).strip(), str(fiber_vendor).strip(), str(contact_numbers).strip() if contact_numbers else None, str(email_address).strip() if email_address else None, str(route_details).strip() if route_details else None))

        conn.commit()
        return {"status": "success", "message": "Fiber record saved successfully."}

    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database save failure: {str(e)}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

##User_Auth

@app.get("/api/admin/users/all")
async def api_get_all_users(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin authorization required.")
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    # FIXED: Added email_id, employee_id, and is_first_login to the SELECT query
    cursor.execute("""
        SELECT id, username, email_id, full_name, employee_id, role, is_first_login 
        FROM users 
        ORDER BY id ASC
    """)
    user_records = cursor.fetchall()
    cursor.close()
    conn.close()
    return user_records

@app.post("/api/admin/users/update")
async def api_update_user_profile(payload: UserUpdateModel, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin authorization required.")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if payload.password and payload.password.strip():
            new_hash = hash_password(payload.password.strip())
            cursor.execute(
                """UPDATE users 
                   SET username = %s, email_id = %s, full_name = %s, employee_id = %s, role = %s, password_hash = %s 
                   WHERE id = %s""",
                (payload.username.strip().lower(), payload.email_id.strip().lower(), payload.full_name.strip(), payload.employee_id.strip(), payload.role, new_hash, payload.user_id)
            )
        else:
            cursor.execute(
                """UPDATE users 
                   SET username = %s, email_id = %s, full_name = %s, employee_id = %s, role = %s 
                   WHERE id = %s""",
                (payload.username.strip().lower(), payload.email_id.strip().lower(), payload.full_name.strip(), payload.employee_id.strip(), payload.role, payload.user_id)
            )
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Database constraint violation: {str(e)}")
    finally:
        cursor.close(); conn.close()

@app.post("/api/admin/users/create")
async def api_create_new_user(payload: UserCreateModel, user=Depends(get_current_user)):
    # Strict Admin Security Authorization Guard
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin authorization required.")
    
    if not payload.password or len(payload.password.strip()) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters long.")

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Check if username or email already exists
        cursor.execute("SELECT id FROM users WHERE LOWER(username) = LOWER(%s) OR LOWER(email_id) = LOWER(%s)", 
                       (payload.username.strip(), payload.email_id.strip()))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Username or Email ID already registered within system inventory.")

        # Hash the admin-defined password securely
        new_password_hash = hash_password(payload.password.strip())
        
        # Insert the metadata layout mapping your specifications
        query = """
            INSERT INTO users (username, email_id, full_name, employee_id, role, password_hash)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (
            payload.username.strip().lower(),
            payload.email_id.strip().lower(),
            payload.full_name.strip(),
            payload.employee_id.strip(),
            payload.role,
            new_password_hash
        ))
        conn.commit()
        return {"status": "success", "message": f"User account tracking profile created successfully for {payload.username}."}
    except HTTPException as http_e:
        raise http_e
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Database target mapping constraint dropped: {str(e)}")
    finally:
        cursor.close()
        conn.close()

@app.post("/api/admin/users/delete/{target_id}")
async def api_delete_user(target_id: int, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin authorization required.")
        
    if user["id"] == target_id:
        raise HTTPException(status_code=400, detail="Administrative Safeguard: You cannot delete your own active session account.")

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM users WHERE id = %s", (target_id,))
        conn.commit()
        return {"status": "success", "message": "User profile successfully removed from core systems inventory."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database execution layer failure: {str(e)}")
    finally:
        cursor.close(); conn.close()

@app.post("/api/provision/welcome-handover")
async def api_welcome_handover(
    background_tasks: BackgroundTasks,
    customer_name: str = Form(...),
    circuit_id: str = Form(...),
    bandwidth_speed: str = Form(...),
    commissioning_date: str = Form(...),
    wan_ip_details: str = Form(...),
    usable_ips: str = Form(...),
    default_gateway: str = Form(...),
    subnet_mask: str = Form(...),
    customer_email: str = Form(...),
    cc_emails: str = Form(""),
    testing_snap: UploadFile = File(...),
    escalation_file: UploadFile = File(...), 
    user=Depends(get_current_user)
):
    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    msg['To'] = customer_email.strip()
    
    recipients_cc = list(GLOBAL_MANDATORY_CC)
    if cc_emails:
        custom_emails = [email.strip() for email in cc_emails.split(",") if email.strip()]
        for c_email in custom_emails:
            if c_email not in recipients_cc:
                recipients_cc.append(c_email)
    msg['Cc'] = ", ".join(recipients_cc)
    msg['Subject'] = f"Welcome to TeleGlobal Communications || Link Delivery Handover - {circuit_id}"
    
    html_template = f"""<!DOCTYPE html>
<html lang="en">
<body style="background-color:#f4f7fa;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#f4f7fa">
    <tr>
        <td align="center" style="padding:30px 15px;">
            <table width="700" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:8px;">
                <tr>
                    <td bgcolor="#0b4d91" align="center" style="padding:35px 25px;">
                        <h2 style="margin:0; color:#ffffff;">Link Delivery Confirmation</h2>
                    </td>
                </tr>
                <tr>
                    <td style="padding:40px 35px; color:#333333; font-size:14px; line-height:24px;">
                        <p><strong>Dear Sir,</strong></p>
                        <p>Your <strong>{bandwidth_speed} Internet Leased Line</strong> has been successfully commissioned.</p>
                        <table width="100%" cellpadding="12" cellspacing="0" border="1" style="border-collapse:collapse; margin-top:25px;">
                            <tr bgcolor="#f7f9fc"><td>Circuit ID</td><td><strong>{circuit_id}</strong></td></tr>
                            <tr><td>Customer Name</td><td>{customer_name}</td></tr>
                            <tr bgcolor="#f7f9fc"><td>Bandwidth Speed</td><td>{bandwidth_speed}</td></tr>
                            <tr><td>Commissioning Date</td><td>{commissioning_date}</td></tr>
                            <tr bgcolor="#f7f9fc"><td>WAN IP Details</td><td>Usable Range: <strong>{usable_ips}</strong> ({wan_ip_details})</td></tr>
                            <tr><td>Default Gateway</td><td>{default_gateway}</td></tr>
                            <tr bgcolor="#f7f9fc"><td>Subnet Mask</td><td>{subnet_mask}</td></tr>
                        </table>
                    </td>
                </tr>
            </table>
        </td>
    </tr>
</table>
</body>
</html>"""
    
    msg.attach(MIMEText(append_signature(html_template, user), 'html'))
    
    if testing_snap and testing_snap.filename:
        try:
            snap_bytes = await testing_snap.read()
            if len(snap_bytes) > 0:
                part1 = MIMEBase('application', 'octet-stream')
                part1.set_payload(snap_bytes)
                encoders.encode_base64(part1)
                part1.add_header('Content-Disposition', f'attachment; filename="{testing_snap.filename}"')
                msg.attach(part1)
        except Exception as e:
            print(f"Testing Snap processing error: {str(e)}")

    if escalation_file and escalation_file.filename:
        try:
            esc_bytes = await escalation_file.read()
            if len(esc_bytes) > 0:
                part2 = MIMEBase('application', 'octet-stream')
                part2.set_payload(esc_bytes)
                encoders.encode_base64(part2)
                part2.add_header('Content-Disposition', f'attachment; filename="{escalation_file.filename}"')
                msg.attach(part2)
        except Exception as e:
            print(f"Escalation Matrix File processing error: {str(e)}")

    all_recipients = [customer_email.strip()] + recipients_cc
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), all_recipients)
    return {"status": "success", "message": "Welcome Onboarding Pack dispatched successfully."}

# --- SCALABLE ALARM CONFIGURATION MATRIX ---

def determine_email_template(issue_category: str, status: str) -> str:
    if not issue_category:
        cat = ""
    else:
        cat = issue_category.strip().lower()

    # APPEND THIS MATCHING LOGIC FOR CUSTOM MAIL
    if cat in ["custom mail", "custom mail dispatch", "custom notification"]:
        return "custom_mail.html"

    # Existing template conditionals
    if cat in ["primary default update", "primary default", "primary", "general issue"]:
        return "primary_up.html" if status == "Closed" else "primary.html"
    elif cat == "ip is not pingable":
        return "ip_up.html" if status == "Closed" else "ip_down.html"
    elif cat in ["switch isolated", "switch is isolated"]:
        return "switch_up.html" if status == "Closed" else "switch_down.html"
    elif cat in ["media device down", "media device"]:
        return "mkt_up.html" if status == "Closed" else "mkt.html"
    elif cat == "high bandwidth utilization":
        return "high.html"

    if status == "Closed":
        return "link_up.html"
    elif status == "In Monitoring":
        return "link_monitoring.html"
    else:
        return "link_down.html"

def assign_priority_and_sla(issue_category: str):
    now = datetime.utcnow()
    
    if issue_category in ["Core Backbone Down", "Complete Site Isolation", "Total Blackout"]:
        return "P1", now + timedelta(hours=1)
    elif issue_category in ["high bandwidth utilization", "High Packet Loss", "Latency Issues", "Redundancy Lost"]:
        return "P2", now + timedelta(hours=4)
    elif issue_category in ["Interface Flapping", "Minor Config Error"]:
        return "P3", now + timedelta(hours=24)
    else:
        return "P4", now + timedelta(days=3)

@app.post("/api/ticket/raise")
async def process_raise_ticket(
    background_tasks: BackgroundTasks,
    circuit_id: str = Form(...),
    issue_category: str = Form(...),
    root_cause_segment: str = Form(...),
    status: str = Form(...),
    assigned_team: str = Form(...),
    generate_ticket: str = Form("true"),
    cc_emails: str = Form(""),
    attachment: UploadFile = File(None),
    user=Depends(get_current_user)
):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cursor.execute("SELECT * FROM customers WHERE LOWER(TRIM(circuit_id)) = LOWER(%s)", (circuit_id.strip(),))
        customer = cursor.fetchone()
        if not customer:
            raise HTTPException(status_code=400, detail="Cannot log ticket against unverified Circuit.")

        engineer_identity = user["full_name"]
        if "|" in engineer_identity:
            engineer_identity = engineer_identity.split("|")[0].strip()

        if generate_ticket == "true":
            closed_at_timestamp = datetime.now() if status == "Closed" else None
            closed_by_identity = engineer_identity if status == "Closed" else None

            cursor.execute(
                """INSERT INTO tickets (circuit_id, issue_category, root_cause_segment, status, assigned_team, open_by_name, closed_by_name, created_at, closed_at, resolution_minutes)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, %s, 0) RETURNING ticket_id, created_at""",
                (circuit_id.strip(), issue_category, root_cause_segment, status, assigned_team, engineer_identity, closed_by_identity, closed_at_timestamp)
            )
            inserted_row = cursor.fetchone()
            ticket_id = inserted_row['ticket_id']
            
            if status == "Closed":
                time_delta = closed_at_timestamp - inserted_row['created_at'].replace(tzinfo=None)
                duration_minutes = max(1, int(time_delta.total_seconds() / 60))
                cursor.execute("UPDATE tickets SET resolution_minutes = %s WHERE ticket_id = %s", (duration_minutes, ticket_id))

            # --- DAILY RESET TICKET ID LOGIC ---
            cursor.execute(
                "SELECT COUNT(*) as daily_count FROM tickets WHERE created_at::date = CURRENT_DATE"
            )
            daily_count = cursor.fetchone()['daily_count']

            conn.commit()
            formatted_ticket_id = f"TCPL{inserted_row['created_at'].strftime('%d%m%y')}{daily_count:02d}"
        else:
            # --- DIRECT EMAIL MODE FIXED LOGIC ---
            # Generate uniform TCPL reference ID using today's total ticket count sequence
            now = datetime.now()
            cursor.execute("SELECT COUNT(*) as count FROM tickets WHERE created_at::date = CURRENT_DATE")
            count_today = cursor.fetchone()['count'] + 1
            formatted_ticket_id = f"TCPL{now.strftime('%d%m%y')}{count_today:02d}"

    except Exception as db_err:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database Logging Error: {str(db_err)}")
    finally:
        cursor.close()
        conn.close()

    # --- EMAIL PREPARATION & INLINE ATTACHMENT HANDLING ---
    recipients_cc = list(GLOBAL_MANDATORY_CC)
    if cc_emails:
        custom_emails = [email.strip() for email in cc_emails.split(",") if email.strip()]
        for c_email in custom_emails:
            if c_email not in recipients_cc:
                recipients_cc.append(c_email)

    # Determine Customer & Company Name
    resolved_customer_name = customer.get("customer_name") if customer and customer.get("customer_name") else "Valued Client"
    resolved_company_name = customer.get("company_name") or customer.get("customer_name") or "Valued Client"

    # Read attached file bytes (if provided)
    attachment_bytes = None
    attachment_filename = None
    if attachment and attachment.filename:
        attachment_bytes = await attachment.read()
        attachment_filename = attachment.filename

    # Determine template
    template_file = determine_email_template(issue_category, status)

    # Build inline HTML image block if snapshot is provided for High Bandwidth Utilization
    inline_snapshot_html = ""
    if attachment_bytes and template_file == "high.html":
        inline_snapshot_html = """
        <div style="margin:25px 0;padding:18px;background:#fff8e6;border-left:5px solid #d97706;border-radius:4px;text-align:left;">
            <b style="display:block;margin-bottom:12px;color:#b45309;font-size:15px;font-weight:700;">Current Link Utilization Snapshot:</b>
            <div style="text-align:center;padding:5px 0;">
                <img src="cid:utilization_snapshot_img" alt="Bandwidth Utilization Snapshot" style="max-width:100%;height:auto;border:1px solid #d9e2ec;border-radius:4px;display:inline-block;margin:0 auto;box-shadow:0 2px 6px rgba(0,0,0,0.06);">
            </div>
        </div>
        """

    # Load HTML template file
    try:
        with open(f"/opt/noc-app/templates/emails/{template_file}", "r", encoding="utf-8") as html_file:
            html_template_data = html_file.read()

        final_html_body = html_template_data\
            .replace("{customer_name}", str(resolved_customer_name))\
            .replace("{circuit_id}", str(circuit_id))\
            .replace("{{IP_ADDRESS}}", str(circuit_id))\
            .replace("{ticket_id}", str(formatted_ticket_id))\
            .replace("{operator_name}", str(engineer_identity))\
            .replace("{status}", str(status))\
            .replace("{issue_category}", str(issue_category))\
            .replace("{issue_subject}", str(issue_category))\
            .replace("{etr}", "2 Hours")\
            .replace("{root_cause_segment}", str(root_cause_segment))\
            .replace("{assigned_team}", str(assigned_team))\
            .replace("{remark_note}", "Ticket Initialization.")\
            .replace("{SNAPSHOT_CONTAINER}", inline_snapshot_html)

        email_content_with_signature = append_signature(final_html_body, user)
    except Exception:
        # Fallback text if template fails
        email_content_with_signature = f"Dear Operations Team,\n\nIncident Reference: #{formatted_ticket_id}\nCircuit Reference: {circuit_id}\nRegards,\n{engineer_identity}"

    # Construct and dispatch MIME email message
    if customer and customer.get("customer_email"):
        # Updated Subject Line including Company Name
        subject_line = f"[NOC Ticket #{formatted_ticket_id}] {issue_category} | {resolved_company_name} | Circuit ID: {circuit_id}"
        
        msg = build_email_message(
            sender=SMTP_USER,
            to_email=customer["customer_email"],
            cc_list=recipients_cc,
            subject=subject_line,
            html_body=email_content_with_signature,
            file_bytes=attachment_bytes,
            filename=attachment_filename,
            is_high_utilization=(template_file == "high.html")
        )

        all_recipients = [customer["customer_email"]] + recipients_cc
        background_tasks.add_task(send_smtp_email_background, msg.as_string(), all_recipients)

    return {"status": "success", "ticket_id": formatted_ticket_id}

@app.post("/api/ticket/update-status")
async def update_ticket_status(payload: dict, background_tasks: BackgroundTasks, user=Depends(get_current_user)):
    ticket_id = payload.get("ticket_id")
    target_status = payload.get("status")
    remark_note = payload.get("remark_note", "")
    clean_remark = remark_note.strip() if remark_note.strip() else "No additional comments provided."

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cursor.execute("SELECT circuit_id, assigned_team, open_by_name, created_at, issue_category FROM tickets WHERE ticket_id = %s", (ticket_id,))
        ticket_meta = cursor.fetchone()
        if not ticket_meta:
            raise HTTPException(status_code=404, detail="Ticket record not found.")

        engineer_identity = user["full_name"]
        if "|" in engineer_identity:
            engineer_identity = engineer_identity.split("|")[0].strip()
        
        if target_status == "Closed":
            closed_at = datetime.now(timezone.utc)
            created_at_tz = ticket_meta['created_at'].replace(tzinfo=timezone.utc)
            time_delta = closed_at - created_at_tz
            duration_minutes = max(1, int(time_delta.total_seconds() / 60))
            cursor.execute(
                "UPDATE tickets SET status = %s, closed_by_name = %s, closed_at = %s, resolution_minutes = %s WHERE ticket_id = %s",
                (target_status, engineer_identity, closed_at, duration_minutes, ticket_id)
            )
        else:
            cursor.execute("UPDATE tickets SET status = %s, closed_by_name = NULL, closed_at = NULL, resolution_minutes = 0 WHERE ticket_id = %s", (target_status, ticket_id))
            
        cursor.execute("SELECT customer_name, company_name, customer_email FROM customers WHERE LOWER(TRIM(circuit_id)) = LOWER(TRIM(%s))", (ticket_meta["circuit_id"],))
        customer_meta = cursor.fetchone()
        conn.commit()
    except Exception as db_err:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database operational failure: {str(db_err)}")
    finally:
        cursor.close()
        conn.close()

    # --- SUPPRESS EMAIL NOTIFICATION FOR HIGH BANDWIDTH UTILIZATION ISSUES ---
    if ticket_meta and ticket_meta.get("issue_category") == "High Bandwidth Utilization":
        return {
            "status": "success",
            "message": "Ticket status updated successfully in database. Email notification bypassed for High Bandwidth Utilization."
        }

    # --- EMAIL DISPATCH FOR OTHER ISSUE CATEGORIES ---
    template_file = determine_email_template(ticket_meta.get("issue_category"), target_status)

    try:
        with open(f"/opt/noc-app/templates/emails/{template_file}", "r", encoding="utf-8") as html_file:
            html_template_data = html_file.read()
    except Exception as io_err:
        raise HTTPException(status_code=500, detail=f"Failed loading HTML template: {str(io_err)}")

    recipients_cc = list(GLOBAL_MANDATORY_CC)
    formatted_ticket_id = f"TCPL{ticket_meta['created_at'].strftime('%d%m%y')}{ticket_id:02d}"
    
    # Resolve Customer Name & Company Name
    resolved_customer_name = customer_meta["customer_name"] if customer_meta and customer_meta.get("customer_name") else "Valued Client"
    resolved_company_name = customer_meta.get("company_name") or resolved_customer_name if customer_meta else "Valued Client"

    final_html_body = html_template_data\
        .replace("{customer_name}", str(resolved_customer_name))\
        .replace("{circuit_id}", str(ticket_meta["circuit_id"]))\
        .replace("{{IP_ADDRESS}}", str(ticket_meta["circuit_id"]))\
        .replace("{ticket_id}", str(formatted_ticket_id))\
        .replace("{operator_name}", str(engineer_identity))\
        .replace("{status}", str(target_status))\
        .replace("{issue_category}", str(ticket_meta.get("issue_category")))\
        .replace("{issue_subject}", str(ticket_meta.get("issue_category")))\
        .replace("{etr}", "Resolved" if target_status == "Closed" else "2 Hours")\
        .replace("{remark_note}", str(clean_remark))

    if customer_meta and customer_meta.get("customer_email"):
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = customer_meta["customer_email"]
        msg['Cc'] = ", ".join(recipients_cc)
        
        # Updated Subject Line with Company Name and Ticket ID
        msg['Subject'] = f"[NOC Ticket #{formatted_ticket_id}] {resolved_company_name} | Internet Link Status Notice [{target_status}] - Circuit ID: {ticket_meta['circuit_id']}"
        
        msg.attach(MIMEText(append_signature(final_html_body, user), 'html'))
        background_tasks.add_task(send_smtp_email_background, msg.as_string(), [customer_meta["customer_email"]] + recipients_cc)

    return {"status": "success"}

@app.get("/api/tickets/recent")
async def get_recent_tickets(
    limit: int = 50, 
    search: Optional[str] = "", 
    status: Optional[str] = "",
    current_user=Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = """
            SELECT 
                ticket_id,
                ticket_id AS raw_ticket_id,
                circuit_id,
                issue_category,
                open_by_name,
                COALESCE(assigned_team, 'Unassigned') AS assigned_team,
                COALESCE(assigned_team, 'Unassigned') AS assigned_operator_team,
                COALESCE(assigned_to_name, open_by_name) AS assigned_to_name,
                closed_by_name,
                status,
                resolution_minutes,
                forwarding_remarks,
                created_at,
                closed_at
            FROM tickets
            WHERE 1=1
        """
        params = []
        
        if status:
            query += " AND status = %s"
            params.append(status)
            
        if search:
            query += " AND (ticket_id::text ILIKE %s OR circuit_id ILIKE %s OR issue_category ILIKE %s)"
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param])
            
        query += " ORDER BY ticket_id DESC LIMIT %s"
        params.append(limit)
        
        cursor.execute(query, tuple(params))
        tickets = cursor.fetchall()
        return tickets
    except Exception as e:
        print(f"[ERROR] /api/tickets/recent: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

from fastapi import Depends, FastAPI, HTTPException
from fastapi.responses import Response
# OpenPyXL Imports for Professional Excel Formatting
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================================================================
# HELPER: DATETIME FORMATTER
# =========================================================================
def format_excel_datetime(dt_val, date_only=False):
    """Safely formats datetime objects or raw DB strings into clean, readable strings."""
    if not dt_val:
        return "N/A"
    if isinstance(dt_val, datetime):
        return dt_val.strftime("%Y-%m-%d %H:%M:%S") if not date_only else dt_val.strftime("%Y-%m-%d")
    
    str_val = str(dt_val).replace('T', ' ').split('.')[0]
    return str_val

# =========================================================================
# OPERATIONAL COMPLIANCE EXPORT DATA AUDIT ENGINE
# =========================================================================
@app.post("/api/reports/download")
async def stream_excel_report_dataset(payload: ReportPayload, user: dict = Depends(get_current_user)):
    report_type = payload.report_type
    start_date = payload.start_date
    end_date = payload.end_date
    team = payload.team

    operator_name = user.get('full_name', 'System Desk')
    now_dt = datetime.now()
    computed_filename = f"teleglobal_{report_type}_{now_dt.strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Initialize Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "NOC Compliance Report"

    # Ensure Grid Lines are visible in Excel
    ws.views.sheetView[0].showGridLines = True

    # Style Definitions
    title_font = Font(name="Calibri", size=15, bold=True, color="003B8E")
    subtitle_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
    meta_font = Font(name="Calibri", size=10, italic=True, color="475569")
    
    section_font = Font(name="Calibri", size=11, bold=True, color="003B8E")
    summary_label_font = Font(name="Calibri", size=10, bold=True, color="334155")
    summary_val_font = Font(name="Calibri", size=11, bold=True, color="003B8E")
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003B8E", end_color="003B8E", fill_type="solid")
    
    data_font = Font(name="Calibri", size=10, color="1E293B")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    status_fills = {
        "CLOSED": PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid"),  # Soft Green
        "OPEN": PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),    # Soft Red
        "ACTIVE": PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")   # Soft Yellow
    }

    # Extract Data & Compute Metrics First to Determine Total Column Count
    headers = []
    data_rows = []
    
    # Summary Metrics Tracking
    total_records = 0
    closed_tickets_count = 0
    active_tickets_count = 0
    total_res_minutes = 0

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
    except Exception:
        raise HTTPException(status_code=500, detail="Unable to connect to database engine.")

    try:
        if report_type == "tickets":
            headers = ["Ticket ID", "Circuit ID", "Alarm Category", "Assigned Team", "Fault Location Segment", "Status", "Opened By", "Closed By", "Opened Time", "Closed Time", "Duration Log"]
            
            query = "SELECT * FROM tickets WHERE 1=1"
            params = []
            if start_date:
                query += " AND created_at >= %s"; params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND created_at <= %s"; params.append(f"{end_date} 23:59:59")
            if team and team != "All":
                query += " AND assigned_team = %s"; params.append(team)
            query += " ORDER BY ticket_id DESC"
            
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            
            total_records = len(records)
            for item in records:
                status = item.get('status', '')
                res_mins = item.get('resolution_minutes') or 0
                
                if str(status).lower() == "closed":
                    closed_tickets_count += 1
                    total_res_minutes += res_mins
                else:
                    active_tickets_count += 1

                duration = f"{res_mins} Mins" if str(status).lower() == 'closed' and res_mins else 'Active'
                
                data_rows.append([
                    f"#{item.get('ticket_id')}",
                    item.get('circuit_id', ''),
                    item.get('issue_category', ''),
                    item.get('assigned_team', ''),
                    item.get('root_cause_segment', ''),
                    status,
                    item.get('open_by_name', ''),
                    item.get('closed_by_name', ''),
                    format_excel_datetime(item.get('created_at')),
                    format_excel_datetime(item.get('closed_at')),
                    duration
                ])

        elif report_type == "welcome_links":
            headers = ["Circuit ID", "Customer Name", "Company Name", "Contact Phone / Email", "Site Location Address", "Dispatched Timestamp"]
            query = "SELECT c.circuit_id, c.customer_name, c.company_name, c.customer_email, c.address, wl.dispatched_at FROM customers c INNER JOIN welcome_mail_logs wl ON c.circuit_id = wl.circuit_id WHERE 1=1"
            params = []
            if start_date:
                query += " AND wl.dispatched_at >= %s"; params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND wl.dispatched_at <= %s"; params.append(f"{end_date} 23:59:59")
            query += " ORDER BY wl.dispatched_at DESC"
            
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            total_records = len(records)
            for item in records:
                data_rows.append([
                    item.get('circuit_id', ''),
                    item.get('customer_name', ''),
                    item.get('company_name', 'Teleglobal Client'),
                    item.get('customer_email', ''),
                    item.get('address', 'NOC Managed Location'),
                    format_excel_datetime(item.get('dispatched_at'))
                ])

        elif report_type == "bandwidth_changes":
            headers = ["Circuit ID", "Customer Name", "Company Name", "Old Bandwidth", "New Bandwidth", "Upgradation Date"]
            query = "SELECT c.customer_name, c.company_name, bl.circuit_id, bl.old_bandwidth, bl.new_bandwidth, bl.upgraded_at FROM bandwidth_upgrade_logs bl LEFT JOIN customers c ON bl.circuit_id = c.circuit_id WHERE 1=1"
            params = []
            if start_date:
                query += " AND bl.upgraded_at >= %s"; params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND bl.upgraded_at <= %s"; params.append(f"{end_date} 23:59:59")
            query += " ORDER BY bl.upgraded_at DESC"
            
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            
            if not records:
                fb_query = "SELECT c.customer_name, c.company_name, t.circuit_id, 'Current/Standard' as old_bandwidth, t.root_cause_segment as new_bandwidth, t.created_at as upgraded_at FROM tickets t INNER JOIN customers c ON t.circuit_id = c.circuit_id WHERE t.issue_category ILIKE %s"
                fb_params = ["%bandwidth%"]
                if start_date:
                    fb_query += " AND t.created_at >= %s"; fb_params.append(f"{start_date} 00:00:00")
                if end_date:
                    fb_query += " AND t.created_at <= %s"; fb_params.append(f"{end_date} 23:59:59")
                cursor.execute(fb_query, tuple(fb_params))
                records = cursor.fetchall()
            
            total_records = len(records)
            for item in records:
                data_rows.append([
                    item.get('circuit_id'),
                    item.get('customer_name'),
                    item.get('company_name', 'Teleglobal Client'),
                    item.get('old_bandwidth', 'N/A'),
                    item.get('new_bandwidth', 'N/A'),
                    format_excel_datetime(item.get('upgraded_at'))
                ])

        elif report_type == "link_terminations":
            headers = ["Circuit ID", "Customer Name", "Company Name", "Termination Reason", "Termination Date"]
            query = "SELECT c.customer_name, c.company_name, tl.circuit_id, tl.reason, tl.terminated_at FROM link_termination_logs tl LEFT JOIN customers c ON tl.circuit_id = c.circuit_id WHERE 1=1"
            params = []
            if start_date:
                query += " AND tl.terminated_at >= %s"; params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND tl.terminated_at <= %s"; params.append(f"{end_date} 23:59:59")
            query += " ORDER BY tl.terminated_at DESC"
            
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            
            if not records:
                fb_query = "SELECT c.customer_name, c.company_name, t.circuit_id, t.root_cause_segment as reason, t.created_at as terminated_at FROM tickets t INNER JOIN customers c ON t.circuit_id = c.circuit_id WHERE (t.issue_category ILIKE %s OR t.issue_category ILIKE %s)"
                fb_params = ["%terminat%", "%decommission%"]
                if start_date:
                    fb_query += " AND t.created_at >= %s"; fb_params.append(f"{start_date} 00:00:00")
                if end_date:
                    fb_query += " AND t.created_at <= %s"; fb_params.append(f"{end_date} 23:59:59")
                cursor.execute(fb_query, tuple(fb_params))
                records = cursor.fetchall()
            
            total_records = len(records)
            for item in records:
                data_rows.append([
                    item.get('circuit_id'),
                    item.get('customer_name'),
                    item.get('company_name', 'Teleglobal Client'),
                    item.get('reason', 'Decommission Request Issued'),
                    format_excel_datetime(item.get('terminated_at'))
                ])

    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Database extraction fault: {str(err)}")
    finally:
        cursor.close(); conn.close()

    # 1. Header Banner (Centered & Merged across all active report columns)
    num_cols = len(headers) if headers else 11

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)

    ws['A1'] = "TELEGLOBAL COMMUNICATIONS PVT. LTD."
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws['A2'] = "NOC AUTOMATION ENGINE - COMPLIANCE EXPORT DATA AUDIT"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws['A3'] = f"Generated By: {operator_name} | Date: {now_dt.strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = meta_font
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

    # Row Height formatting for clear header separation
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    # 2. Executive Summary Block
    ws['A5'] = "📊 EXECUTIVE SUMMARY"
    ws['A5'].font = section_font

    avg_res_time = f"{int(total_res_minutes / closed_tickets_count)} Mins" if closed_tickets_count > 0 else "N/A"

    summary_cards = [
        ("Total Records Logged", total_records),
        ("Closed Tickets", closed_tickets_count if report_type == "tickets" else "N/A"),
        ("Active / Pending", active_tickets_count if report_type == "tickets" else "N/A"),
        ("Avg Resolution Time", avg_res_time if report_type == "tickets" else "N/A")
    ]

    # Render Executive Summary Cards
    col_pos = 1
    for label, val in summary_cards:
        # Label cell
        lbl_cell = ws.cell(row=6, column=col_pos, value=label)
        lbl_cell.font = summary_label_font
        lbl_cell.fill = card_fill
        lbl_cell.border = thin_border
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Value cell
        val_cell = ws.cell(row=7, column=col_pos, value=val)
        val_cell.font = summary_val_font
        val_cell.fill = card_fill
        val_cell.border = thin_border
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        col_pos += 1

    # 3. Main Data Table Headers
    header_row_idx = 10
    ws.row_dimensions[header_row_idx].height = 26

    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. Render Table Data Rows
    start_data_row = 11
    for r_idx, row_values in enumerate(data_rows, start=start_data_row):
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            # Formatting & Alignments
            val_str = str(val).upper().strip() if val else ""
            header_name = headers[c_idx - 1]
            
            if c_idx == 1 or "ID" in header_name:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "Status" in header_name:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val_str in status_fills:
                    cell.fill = status_fills[val_str]
            elif "Time" in header_name or "Date" in header_name or "Timestamp" in header_name:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 5. Dynamic Auto-Fit Columns Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            # Skip checking lengths for Top Header rows (A1-A8) to avoid ultra-wide columns
            if cell.row < 10:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # Output Stream Response
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    return Response(
        content=output_stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{computed_filename}\"",
            "Cache-Control": "no-cache"
        }
    )

# =========================================================================
# APPEND THESE ENDPOINTS TO THE BOTTOM OF MAIN.PY
# =========================================================================

@app.get("/api/ticket/search-details/{query_str}")
async def api_search_ticket_details(query_str: str, user=Depends(get_current_user)):
    """ Queries operational logs by formatted ticket reference numbers, raw IDs, or circuit IDs accurately """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    term = query_str.strip()
    wildcard_term = f"%{term}%"
    
    # SQL query dynamically reconstructs standard ticket format 'TCPL' + 'DDMMYY' + 'XX' for comparison
    query = """
        WITH daily_tickets AS (
            SELECT ticket_id, circuit_id, issue_category, root_cause_segment, status, 
                   assigned_team, open_by_name, closed_by_name, created_at, closed_at, resolution_minutes,
                   ROW_NUMBER() OVER (PARTITION BY created_at::date ORDER BY ticket_id ASC) as daily_seq
            FROM tickets
        )
        SELECT ticket_id, circuit_id, issue_category, root_cause_segment, status, assigned_team, open_by_name,
               created_at,
               COALESCE(closed_by_name, '--') as closed_by_name,
               TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI:SS') as opened_at,
               COALESCE(TO_CHAR(closed_at, 'YYYY-MM-DD HH24:MI:SS'), '--') as closed_at,
               resolution_minutes,
               'TCPL' || TO_CHAR(created_at, 'DDMMYY') || LPAD(CAST(daily_seq AS TEXT), 2, '0') as custom_ticket_id
        FROM daily_tickets 
        WHERE LOWER('TCPL' || TO_CHAR(created_at, 'DDMMYY') || LPAD(CAST(daily_seq AS TEXT), 2, '0')) LIKE LOWER(%s)
           OR CAST(ticket_id AS TEXT) = %s
           OR LOWER(circuit_id) LIKE LOWER(%s)
        ORDER BY ticket_id DESC 
        LIMIT 5
    """
    
    try:
        cursor.execute(query, (wildcard_term, term, wildcard_term))
        records = cursor.fetchall()
        
        formatted = []
        for row in records:
            # Structurally generate standardized ticket tracking sequence matching read_recent_tickets
            #ticket_date = row['created_at']
            #custom_ticket_id = f"TCPL{ticket_date.strftime('%d%m%y')}{row['ticket_id']:02d}"
            
            formatted.append({
                "ticket_id": row['custom_ticket_id'],
                "raw_ticket_id": row['ticket_id'],
                "circuit_id": row['circuit_id'],
                "issue_category": row['issue_category'],
                "root_cause_segment": row['root_cause_segment'],
                "status": row['status'],
                "opened_at": row['opened_at'],
                "closed_at": row['closed_at'],
                "resolution_minutes": row['resolution_minutes']
            })
        return formatted
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database operational failure: {str(e)}")
    finally:
        cursor.close()
        conn.close()

from jinja2 import Template

# =========================================================================
# REASON FOR OUTAGE (RFO) AUTOMATION EMAIL DISPATCH PIPELINE
# =========================================================================

@app.post("/api/tools/send-rfo-mail")
async def api_send_rfo_mail(
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_current_user)
):
    """ Dynamically extracts all web form parameters and bridges them into templates/emails/RFO.html """
    try:
        form_data = await request.form()
        payload = dict(form_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse form parameters: {str(e)}")

    # Extract required target email parameters
    customer_email = payload.get("customer_email", "").strip()
    cc_emails = payload.get("cc_emails", "").strip()
    
    if not customer_email:
        raise HTTPException(status_code=400, detail="Missing required field: Client Target Email Address.")

    # Parse and structure recipient arrays
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    recipients_cc = list(GLOBAL_MANDATORY_CC)
    if cc_emails:
        for addr in cc_emails.split(","):
            if addr.strip():
                recipients_cc.append(addr.strip())

    # Map the exact absolute path to your file structure
    template_path = os.path.join("templates", "emails", "RFO.html")
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail=f"Template file not found at path: {template_path}")

    with open(template_path, "r", encoding="utf-8") as f:
        raw_html_content = f.read()

    # Create the context dictionary mapping the web form fields directly to the RFO.html tokens
    email_context = {
        "customer_name": payload.get("customer_name", "").strip(),
        "circuit_id": payload.get("circuit_id", "").strip(),
        "circuit_location": payload.get("circuit_location", "").strip(),
        "bandwidth": payload.get("bandwidth", "").strip(),
        "ticket_id": payload.get("ticket_id", "").strip(),
        "er_type": payload.get("er_type", "").strip(),
        "reference_ticket_id": payload.get("reference_ticket_id", "").strip(),
        "priority": payload.get("priority", "").strip(),
        "creation_time": payload.get("creation_time", "").strip(),
        "resolution_time": payload.get("resolution_time", "").strip(),
        "total_resolution_time": payload.get("total_resolution_time", "").strip(),
        "circuit_up_time": payload.get("circuit_up_time", "").strip(),
        "on_hold_duration": payload.get("on_hold_duration", "").strip(),
        "roopen_case": payload.get("roopen_case", "").strip(),
        "nature_of_fault": payload.get("nature_of_fault", "").strip(),
        "reason_for_outage": payload.get("reason_for_outage", "").strip(),
        "issue_location": payload.get("issue_location", "").strip(),
        "resolution": payload.get("resolution", "").strip(),
        "remarks": payload.get("remarks", "").strip(),
        "current_year": "2026"  # Handled dynamically for your signature footer
    }

    # Render your custom RFO.html file layout using the context above
    jinja_template = Template(raw_html_content)
    html_body = jinja_template.render(**email_context)

    # Compile the final multi-part email envelope configuration
    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    msg['To'] = ", ".join(to_recipients)
    msg['Cc'] = ", ".join(recipients_cc)
    
    company_name = email_context["customer_name"] or "Valued Client"
    circuit_ref = email_context["circuit_id"] or "N/A"
    msg['Subject'] = f"Reason for Outage (RFO) Report || {company_name} || Circuit ID: {circuit_ref}"
    
    msg.attach(MIMEText(append_signature(html_body, user), 'html'))
    all_recipients = to_recipients + recipients_cc

    # Hand off to non-blocking background queue task loop to prevent browser page freeze
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), all_recipients)

    return {"status": "success", "message": f"RFO email compiled and dispatched for ticket {email_context['ticket_id']}"}

# =====================================================================
# EVENT WELCOME MAILER ROUTING SUBSYSTEM
# =====================================================================

# =========================================================================
# EVENT WELCOME MAIL & LOGGING PIPELINE (COMPLETE ARCHITECTURE)
# =========================================================================

# =========================================================================
# EVENT WELCOME MAIL & LOGGING PIPELINE (COMPLETE ARCHITECTURE)
# =========================================================================

@app.get("/system-mail/event-welcome", response_class=HTMLResponse)
async def route_system_mail_event_welcome_page(request: Request, user = Depends(get_optional_user)):
    """
    Renders the secure Event Welcome operational template dashboard interface.
    """
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="system_mail_event_welcome.html", context={"user": user})

@app.get("/api/circuits/search")
async def api_search_event_circuits(query: str = Query(...), user = Depends(get_current_user)):
    """
    Dynamically fetches active inventory profiles matching either Circuit ID or Customer Name.
    """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        search_term = f"%{query.strip()}%"
        db_query = """
            SELECT circuit_id, customer_name, customer_email 
            FROM customers 
            WHERE circuit_id ILIKE %s OR customer_name ILIKE %s
            LIMIT 1;
        """
        cursor.execute(db_query, (search_term, search_term))
        record = cursor.fetchone()
        
        if not record:
            raise HTTPException(status_code=404, detail="No active circuit metrics found matching criteria.")
        return record
    except HTTPException as http_err:
        raise http_err
    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Inventory lookup failure: {str(err)}")
    finally:
        cursor.close()
        conn.close()

# Updated route in main.py
@app.post("/api/tools/send-event-welcome")
async def api_send_event_welcome_mail(
    background_tasks: BackgroundTasks,
    circuit_id: str = Form(...),
    customer_name: str = Form(...),
    bandwidth: str = Form(...),
    event_date: str = Form(...),
    usable_ip: str = Form(...),
    gateway: str = Form(...),
    subnet: str = Form(...),
    customer_email: str = Form(...),
    cc_emails: str = Form(""),
    user = Depends(get_current_user)
):
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))

    # Render the event.html template
    try:
        template = templates.get_template("emails/event.html")
        hydrated_body = template.render({
            "circuit_id": circuit_id.strip(),
            "customer_name": customer_name.strip(),
            "bandwidth": bandwidth.strip(),
            "event_date": event_date.strip(),
            "usable_ip": usable_ip.strip(),
            "gateway": gateway.strip(),
            "subnet": subnet.strip(),
            "operator_name": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(status_code=500, detail=f"Template Error: {str(render_err)}")

    # Construct the message
    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = f"Welcome to Event Connectivity || {customer_name.strip()}"
    
    # Only attach HTML
    msg.attach(MIMEText(append_signature(hydrated_body, user), 'html'))
    
    # Handle Recipients
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip(): cc_list.append(addr.strip())
    
    msg['Cc'] = ", ".join(cc_list)
    recipients.extend(cc_list)

    # Queue background task
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)
    
    return {"status": "success", "message": "Event Welcome Mail dispatched successfully."}

@app.get("/api/reports/download-event-welcome-logs")
async def download_event_welcome_logs(user = Depends(get_current_user)):
    """
    Queries historical operational parameters stored within event_welcome_logs,
    structures the payload into an explicit in-memory stream buffer, and delivers a standard CSV.
    """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        db_query = """
            SELECT id, circuit_id, customer_name, bandwidth, event_date, 
                   usable_ip, gateway, subnet, customer_email, cc_emails, sent_by, sent_at
            FROM event_welcome_logs
            ORDER BY sent_at DESC;
        """
        cursor.execute(db_query)
        records = cursor.fetchall()

        # Initialize the output memory buffers
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        
        # Standardize report header row configurations
        writer.writerow([
            "Log ID", "Circuit ID", "Customer Name", "Bandwidth Speed", "Event Operations Window",
            "Assigned Pool IPs", "Gateway Address", "Subnet Mask Architecture", 
            "Destination Target Email", "Carbon Copy Recipients", "NOC Operator", "Timestamp (Local)"
        ])
        
        # Populate each historical log row configuration entries
        for item in records:
            time_stamp_str = item["sent_at"].strftime("%Y-%m-%d %H:%M:%S") if item["sent_at"] else ""
            writer.writerow([
                item["id"],
                item["circuit_id"],
                item["customer_name"],
                item["bandwidth"],
                item["event_date"],
                item["usable_ip"],
                item["gateway"],
                item["subnet"],
                item["customer_email"],
                item["cc_emails"],
                item["sent_by"],
                time_stamp_str
            ])
        
        csv_buffer.seek(0)
        date_stamp = datetime.now().strftime("%Y%m%d")
        report_filename = f"event_welcome_logs_{date_stamp}.csv"
        
        return StreamingResponse(
            iter([csv_buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={report_filename}"}
        )
    except Exception as query_err:
        raise HTTPException(status_code=500, detail=f"Log compilation processing failed: {str(query_err)}")
    finally:
        cursor.close()
        conn.close()

# Route to handle Access Mail dispatch
@app.post("/api/tools/send-access-mail")
async def api_send_access_mail(
    background_tasks: BackgroundTasks,
    company_name: str = Form(...),
    access_date: str = Form(...),
    purpose: str = Form(...),
    site_location: str = Form(...),
    expected_duration: str = Form(...),
    engineer_name_1: str = Form(""),
    employee_id_1: str = Form(""),
    contact_number_1: str = Form(""),
    id_proof_1: str = Form(""),
    engineer_name_2: str = Form(""),
    employee_id_2: str = Form(""),
    contact_number_2: str = Form(""),
    id_proof_2: str = Form(""),
    engineer_name_3: str = Form(""),
    employee_id_3: str = Form(""),
    contact_number_3: str = Form(""),
    id_proof_3: str = Form(""),
    customer_email: str = Form(...),
    cc_emails: str = Form(""),
    user = Depends(get_current_user)
):
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))

    # Render the emails/access.html template
    try:
        template = templates.get_template("emails/access.html")
        hydrated_body = template.render({
            "company_name": company_name.strip(),
            "access_date": access_date.strip(),
            "purpose": purpose.strip(),
            "site_location": site_location.strip(),
            "expected_duration": expected_duration.strip(),
            "engineer_name_1": engineer_name_1.strip(),
            "employee_id_1": employee_id_1.strip(),
            "contact_number_1": contact_number_1.strip(),
            "id_proof_1": id_proof_1.strip(),
            "engineer_name_2": engineer_name_2.strip(),
            "employee_id_2": employee_id_2.strip(),
            "contact_number_2": contact_number_2.strip(),
            "id_proof_2": id_proof_2.strip(),
            "engineer_name_3": engineer_name_3.strip(),
            "employee_id_3": employee_id_3.strip(),
            "contact_number_3": contact_number_3.strip(),
            "id_proof_3": id_proof_3.strip(),
            "sender_name": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(status_code=500, detail=f"Template Error: {str(render_err)}")

    # Construct the email message
    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = f"Request for Premises Access Approval || {company_name.strip()}"
    
    # Attach HTML body with automated NOC signature
    msg.attach(MIMEText(append_signature(hydrated_body, user), 'html'))
    
    # Handle Recipient Routing
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip(): cc_list.append(addr.strip())
    
    msg['Cc'] = ", ".join(cc_list)
    recipients.extend(cc_list)

    # Queue background sending task
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)
    
    return {"status": "success", "message": "Premises Access Request Mail dispatched successfully."}

# =========================================================================
# BATCH TICKET CLOSURE PIPELINE (NO MAIL / BYPASS MAIL)
# =========================================================================

class BulkCloseRequest(BaseModel):
    ticket_ids: list[int]
    remark_note: Optional[str] = "Bulk Closed without Email Dispatch"

@app.post("/api/ticket/bulk-close-silent")
async def api_bulk_close_tickets_silent(payload: BulkCloseRequest, user=Depends(get_current_user)):
    """
    Closes multiple tickets selected from the UI in bulk.
    Bypasses sending any notification emails to customers.
    """
    if not payload.ticket_ids:
        raise HTTPException(status_code=400, detail="No ticket IDs were provided for bulk closing.")

    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))
    if "|" in engineer_identity:
        engineer_identity = engineer_identity.split("|")[0].strip()

    closed_at = datetime.now(timezone.utc)
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        updated_count = 0
        for ticket_id in payload.ticket_ids:
            # Fetch ticket metadata for duration metrics calculation
            cursor.execute("SELECT created_at FROM tickets WHERE ticket_id = %s", (ticket_id,))
            ticket_meta = cursor.fetchone()
            
            if ticket_meta:
                created_at_tz = ticket_meta['created_at'].replace(tzinfo=timezone.utc)
                time_delta = closed_at - created_at_tz
                duration_minutes = max(1, int(time_delta.total_seconds() / 60))

                cursor.execute(
                    """
                    UPDATE tickets 
                    SET status = 'Closed', 
                        closed_by_name = %s, 
                        closed_at = %s, 
                        resolution_minutes = %s 
                    WHERE ticket_id = %s
                    """,
                    (engineer_identity, closed_at, duration_minutes, ticket_id)
                )
                updated_count += cursor.rowcount

        conn.commit()
        return {
            "status": "success", 
            "message": f"Successfully closed {updated_count} ticket(s) silently without email notification."
        }
    except Exception as db_err:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database execution error: {str(db_err)}")
    finally:
        cursor.close()
        conn.close()

# =========================================================================
# FORCE PASSWORD CHANGE WORKFLOW
# =========================================================================

# Force Password Change Routes
@app.get("/change-password", response_class=HTMLResponse)
async def route_change_password_page(request: Request, user=Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request=request, name="change_password.html", context={"user": user})

@app.post("/api/auth/change-password")
async def api_change_password(payload: ChangePasswordPayload, user=Depends(get_current_user)):
    if len(payload.new_password.strip()) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters long.")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # 1. Verify existing password
        cursor.execute("SELECT password_hash FROM users WHERE id = %s", (user["id"],))
        db_user = cursor.fetchone()
        
        if not db_user or db_user["password_hash"] != hash_password(payload.current_password):
            raise HTTPException(status_code=400, detail="Current password is incorrect.")

        # 2. Update password hash and mark first login as False
        new_hash = hash_password(payload.new_password.strip())
        cursor.execute(
            "UPDATE users SET password_hash = %s, is_first_login = FALSE WHERE id = %s",
            (new_hash, user["id"])
        )
        conn.commit()
        return {"status": "success", "message": "Password successfully updated!"}

    except HTTPException as http_e:
        raise http_e
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database operational error: {str(e)}")
    finally:
        cursor.close()
        conn.close()

@app.post("/api/admin/users/force-password-reset/{target_user_id}")
async def api_admin_force_password_reset(target_user_id: int, user=Depends(get_current_user)):
    # Check if performing user has admin privileges
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin privileges required.")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cursor.execute(
            "UPDATE users SET is_first_login = TRUE WHERE id = %s",
            (target_user_id,)
        )
        conn.commit()
        return {"status": "success", "message": f"User ID {target_user_id} will be forced to change password on next login."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        cursor.close()
        conn.close()

# =========================================================================
# CUSTOM MAIL DISPATCH PIPELINE
# =========================================================================
@app.post("/api/tools/send-custom-mail")
async def api_send_custom_mail(
    background_tasks: BackgroundTasks,
    circuit_id: str = Form(...),
    company_name: str = Form(...),
    customer_email: str = Form(...),
    mail_subject: str = Form(...),
    mail_context: str = Form(...),
    cc_emails: str = Form(""),
    attachment: UploadFile = File(None),
    user=Depends(get_current_user)
):
    """
    Handles dynamic custom email broadcasts, hydration of templates/emails/custom_mail.html,
    and background dispatch to target customer email lists.
    """
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))
    if "|" in engineer_identity:
        engineer_identity = engineer_identity.split("|")[0].strip()

    # Normalize/Format custom text paragraphs to preserve line breaks in HTML
    formatted_context = mail_context.strip().replace("\n", "<br>")

    # Render Template Body
    try:
        template = templates.get_template("emails/custom_mail.html")
        hydrated_body = template.render({
            "circuit_id": circuit_id.strip(),
            "company_name": company_name.strip(),
            "mail_context": formatted_context,
            "operator_name": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(
            status_code=500, 
            detail=f"Failed compiling custom mail template placeholders: {str(render_err)}"
        )

    # Email Dispatch Construction
    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in customer_email.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = mail_subject.strip()

    # Build CC routing list
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip() and addr.strip() not in cc_list:
                cc_list.append(addr.strip())
    
    if cc_list:
        msg['Cc'] = ", ".join(cc_list)
        recipients.extend(cc_list)

    # Attach Hydrated Body with Engineer Signature
    msg.attach(MIMEText(append_signature(hydrated_body, user), 'html'))

    # Handle Attachments
    if attachment and attachment.filename:
        try:
            attachment_bytes = await attachment.read()
            if len(attachment_bytes) > 0:
                part = MIMEApplication(attachment_bytes)
                part['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
                msg.attach(part)
        except Exception as file_err:
            raise HTTPException(
                status_code=500, 
                detail=f"Error parsing custom attachment: {str(file_err)}"
            )

    # Trigger Non-blocking Background Queue Execution
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)

    return {
        "status": "success", 
        "message": f"Custom email dispatched successfully to {customer_email}."
    }

# ==========================================
# DAILY SHIFT UPDATE ROUTES & INFRASTRUCTURE
# ==========================================

# Helper to fetch latest active updates content
def get_latest_shift_update():
    conn = get_db_connection()
    if not conn:
        return ""
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT updates_content FROM daily_shift_updates 
                ORDER BY updated_at DESC LIMIT 1
            """)
            result = cur.fetchone()
            return result["updates_content"] if result else ""
    except Exception as e:
        print(f"Error fetching latest shift update: {e}")
        return ""
    finally:
        conn.close()

# API endpoint to retrieve active updates content dynamically
@app.get("/api/daily-update/latest")
async def api_get_latest_daily_update(user = Depends(get_current_user)):
    content = get_latest_shift_update()
    return {"status": "success", "updates_content": content}

# Route to render the Daily Shift Update UI page
@app.get("/system-mail/update", response_class=HTMLResponse)
async def route_system_mail_update_page(request: Request, user = Depends(get_optional_user)):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Retrieve current active update text to populate text area
    latest_updates = get_latest_shift_update()
    return templates.TemplateResponse(
        request=request, 
        name="system_mail_update.html", 
        context={
            "user": user, 
            "initial_updates": latest_updates,
            "current_date": datetime.now().strftime("%d/%m/%Y")
        }
    )

# Route to handle Daily Shift Update Email dispatch & State Save
from fastapi import BackgroundTasks, Depends, Form, HTTPException


# Route to handle Daily Shift Update Email dispatch & State Save
@app.post("/api/tools/send-daily-update")
async def api_send_daily_update(
    background_tasks: BackgroundTasks,
    shift_date: str = Form(...),          # e.g., "28/07/2026"
    shift_type: str = Form(...),          # MNG, AFT, NGT
    updates_content: str = Form(...),     # Raw text area content
    recipients_to: str = Form(...),       # Default: noc@teleglobal.in
    cc_emails: str = Form(""),
    user = Depends(get_current_user)
):
    engineer_identity = user.get("full_name", user.get("username", "NOC Specialist"))

    # Map Shift Code to full descriptive name
    shift_map = {
        "MNG": "Morning Shift",
        "AFT": "Afternoon Shift",
        "NGT": "Night Shift"
    }
    shift_type_label = shift_map.get(shift_type.strip(), f"{shift_type.strip()} Shift")

    # 1. Save state to DB
    conn = get_db_connection()
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO daily_shift_updates (shift_date, shift_type, updates_content, updated_by)
                    VALUES (%s, %s, %s, %s)
                """, (shift_date.strip(), shift_type.strip(), updates_content.strip(), engineer_identity))
                conn.commit()
        except Exception as db_err:
            print(f"DB Save Error: {db_err}")
        finally:
            conn.close()

    # 2. Convert line breaks into structured list items and remove existing numbers
    raw_lines = [line.strip() for line in updates_content.strip().split("\n") if line.strip()]
    cleaned_items = []
    
    for line in raw_lines:
        # Regex strips leading numbers like "1. ", "2) ", "10 .", "1.1 " to prevent duplicate numbering
        cleaned = re.sub(r'^\s*\d+[\s\.\)]*', '', line).strip()
        if cleaned:
            cleaned_items.append(cleaned)

    # 3. Render emails/daily_update.html template
    try:
        template = templates.get_template("emails/daily_update.html")
        hydrated_body = template.render({
            "shift_date": shift_date.strip(),
            "shift_type": shift_type.strip(),
            "shift_type_label": shift_type_label,
            "update_items": cleaned_items,
            "sender_name": engineer_identity
        })
    except Exception as render_err:
        raise HTTPException(status_code=500, detail=f"Template Error: {str(render_err)}")

    # 4. Construct Email Message
    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    to_recipients = [addr.strip() for addr in recipients_to.split(",") if addr.strip()]
    msg['To'] = ", ".join(to_recipients)
    msg['Subject'] = f"RE: Updates || {shift_date.strip()} || {shift_type.strip()}"

    # Attach HTML body with automated professional NOC signature appended
    final_html_with_signature = append_signature(hydrated_body, user)
    msg.attach(MIMEText(final_html_with_signature, 'html'))

    # Recipients & CC logic
    recipients = list(to_recipients)
    cc_list = list(GLOBAL_MANDATORY_CC)
    if cc_emails.strip():
        for addr in cc_emails.split(","):
            if addr.strip(): 
                cc_list.append(addr.strip())

    msg['Cc'] = ", ".join(cc_list)
    recipients.extend(cc_list)

    # Dispatch via background task
    background_tasks.add_task(send_smtp_email_background, msg.as_string(), recipients)

    return {"status": "success", "message": "Daily Shift Update dispatched and carry-forward state saved!"}

from typing import List, Optional

from fastapi import BackgroundTasks, Depends, HTTPException, Response
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Global Mail Server Settings
SMTP_SERVER = "mail.teleglobal.in"
SMTP_PORT = 465
SMTP_USER = "noc@teleglobal.in"
SMTP_PASSWORD = "8QKti-lme88&"
GLOBAL_MANDATORY_CC = ["noc@teleglobal.in"]

# Load Jinja Templates
email_jinja_env = Environment(
    loader=FileSystemLoader("/opt/noc-app/templates/emails")
)

class SendReportMailPayload(BaseModel):
    report_type: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    team: Optional[str] = None
    recipients: Optional[List[str]] = []
    cc_list: Optional[List[str]] = []

def format_display_date(date_str):
    """Converts YYYY-MM-DD into readable date string (e.g. 23 July 2026)."""
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%d %B %Y")
    except Exception:
        return date_str

def format_excel_datetime(dt_obj):
    if not dt_obj:
        return ""
    if isinstance(dt_obj, datetime):
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")
    return str(dt_obj)

def generate_excel_bytes(payload: SendReportMailPayload, user: dict):
    """Generates Excel File AND returns structured data for HTML rendering."""
    report_type = payload.report_type
    start_date = payload.start_date
    end_date = payload.end_date
    team = payload.team

    operator_name = user.get("full_name", "NOC Desk")
    now_dt = datetime.now()
    filename = (
        f"teleglobal_{report_type}_{now_dt.strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "NOC Compliance Report"
    ws.views.sheetView[0].showGridLines = True

    title_font = Font(name="Calibri", size=14, bold=True, color="003B8E")
    subtitle_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
    meta_font = Font(name="Calibri", size=10, italic=True, color="475569")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="003B8E", end_color="003B8E", fill_type="solid"
    )
    data_font = Font(name="Calibri", size=10, color="1E293B")

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers, data_rows = [], []
    total_records = closed_tickets_count = active_tickets_count = 0

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if report_type == "tickets":
            headers = [
                "Ticket ID",
                "Circuit ID",
                "Alarm Category",
                "Assigned Team",
                "Fault Segment",
                "Status",
                "Opened By",
                "Closed By",
                "Opened Time",
                "Closed Time",
                "Duration",
            ]
            query = "SELECT * FROM tickets WHERE 1=1"
            params = []
            if start_date:
                query += " AND created_at >= %s"
                params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND created_at <= %s"
                params.append(f"{end_date} 23:59:59")
            if team and team != "All":
                query += " AND assigned_team = %s"
                params.append(team)
            query += " ORDER BY ticket_id DESC"
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            total_records = len(records)
            for item in records:
                status = str(item.get("status", ""))
                res_mins = item.get("resolution_minutes") or 0
                if status.lower() == "closed":
                    closed_tickets_count += 1
                else:
                    active_tickets_count += 1

                duration = (
                    f"{res_mins} Mins"
                    if status.lower() == "closed" and res_mins
                    else "Active"
                )
                data_rows.append([
                    f"#{item.get('ticket_id')}",
                    item.get("circuit_id", ""),
                    item.get("issue_category", ""),
                    item.get("assigned_team", ""),
                    item.get("root_cause_segment", ""),
                    status.title(),
                    item.get("open_by_name", ""),
                    item.get("closed_by_name", ""),
                    format_excel_datetime(item.get("created_at")),
                    format_excel_datetime(item.get("closed_at")),
                    duration,
                ])

        elif report_type == "welcome_links":
            headers = [
                "Circuit ID",
                "Customer Name",
                "Company Name",
                "Contact Info",
                "Address",
                "Dispatched At",
            ]
            query = "SELECT c.circuit_id, c.customer_name, c.company_name, c.customer_email, c.address, wl.dispatched_at FROM customers c INNER JOIN welcome_mail_logs wl ON c.circuit_id = wl.circuit_id WHERE 1=1"
            params = []
            if start_date:
                query += " AND wl.dispatched_at >= %s"
                params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND wl.dispatched_at <= %s"
                params.append(f"{end_date} 23:59:59")
            query += " ORDER BY wl.dispatched_at DESC"
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            total_records = len(records)
            for item in records:
                data_rows.append([
                    item.get("circuit_id", ""),
                    item.get("customer_name", ""),
                    item.get("company_name", "Teleglobal Client"),
                    item.get("customer_email", ""),
                    item.get("address", "NOC Location"),
                    format_excel_datetime(item.get("dispatched_at")),
                ])

        elif report_type == "bandwidth_changes":
            headers = [
                "Circuit ID",
                "Customer Name",
                "Company Name",
                "Old Bandwidth",
                "New Bandwidth",
                "Upgraded At",
            ]
            query = "SELECT c.customer_name, c.company_name, bl.circuit_id, bl.old_bandwidth, bl.new_bandwidth, bl.upgraded_at FROM bandwidth_upgrade_logs bl LEFT JOIN customers c ON bl.circuit_id = c.circuit_id WHERE 1=1"
            params = []
            if start_date:
                query += " AND bl.upgraded_at >= %s"
                params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND bl.upgraded_at <= %s"
                params.append(f"{end_date} 23:59:59")
            query += " ORDER BY bl.upgraded_at DESC"
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            total_records = len(records)
            for item in records:
                data_rows.append([
                    item.get("circuit_id"),
                    item.get("customer_name"),
                    item.get("company_name", "Teleglobal Client"),
                    item.get("old_bandwidth", "N/A"),
                    item.get("new_bandwidth", "N/A"),
                    format_excel_datetime(item.get("upgraded_at")),
                ])

        elif report_type == "link_terminations":
            headers = [
                "Circuit ID",
                "Customer Name",
                "Company Name",
                "Reason",
                "Terminated At",
            ]
            query = "SELECT c.customer_name, c.company_name, tl.circuit_id, tl.reason, tl.terminated_at FROM link_termination_logs tl LEFT JOIN customers c ON tl.circuit_id = c.circuit_id WHERE 1=1"
            params = []
            if start_date:
                query += " AND tl.terminated_at >= %s"
                params.append(f"{start_date} 00:00:00")
            if end_date:
                query += " AND tl.terminated_at <= %s"
                params.append(f"{end_date} 23:59:59")
            query += " ORDER BY tl.terminated_at DESC"
            cursor.execute(query, tuple(params))
            records = cursor.fetchall()
            total_records = len(records)
            for item in records:
                data_rows.append([
                    item.get("circuit_id"),
                    item.get("customer_name"),
                    item.get("company_name", "Teleglobal Client"),
                    item.get("reason", "Decommission Request"),
                    format_excel_datetime(item.get("terminated_at")),
                ])
    finally:
        cursor.close()
        conn.close()

    # Excel Layout Formatting
    num_cols = len(headers) if headers else 10
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)

    ws["A1"] = "TELEGLOBAL COMMUNICATIONS PVT. LTD."
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = (
        f"NOC Operational Report | Generated By: {operator_name} | {now_dt.strftime('%Y-%m-%d %H:%M:%S')}"
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 24
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for r_idx, row_values in enumerate(data_rows, start=5):
        ws.row_dimensions[r_idx].height = 19
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Clean 3 Summary Metric Cards (Removed Avg Resolution Time)
    summary_cards = [
        {"label": "Total Logged Records", "val": total_records},
        {
            "label": "Closed Tickets",
            "val": (
                closed_tickets_count if report_type == "tickets" else "N/A"
            ),
        },
        {
            "label": "Active / Pending",
            "val": (
                active_tickets_count if report_type == "tickets" else "N/A"
            ),
        },
    ]

    return (
        output.getvalue(),
        filename,
        headers,
        data_rows,
        summary_cards,
    )

def send_email_in_background(
    recipients, cc_list, subject, html_body, excel_bytes, filename
):
    """Executes SMTP background transmission with robust MIME multipart encoding."""
    try:
        # Create Root MIMEMultipart Container (Mixed mode for attachments)
        msg = MIMEMultipart("mixed")
        msg["From"] = SMTP_USER
        msg["To"] = ", ".join(recipients) if recipients else SMTP_USER
        msg["Cc"] = ", ".join(cc_list)
        msg["Subject"] = subject

        # Attach HTML Body
        html_part = MIMEText(html_body, "html", "utf-8")
        msg.attach(html_part)

        # Attach Excel File with proper MIME type
        excel_part = MIMEApplication(
            excel_bytes,
            _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        excel_part.add_header(
            "Content-Disposition", f'attachment; filename="{filename}"'
        )
        msg.attach(excel_part)

        all_recipients = list(set(recipients + cc_list))

        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_USER, all_recipients, msg.as_string())

    except Exception as e:
        print(f"Background email dispatch error: {str(e)}")

@app.post("/api/reports/send-reports")
async def send_reports_via_email(
    payload: SendReportMailPayload,
    background_tasks: BackgroundTasks,
    user: dict = Depends(get_current_user),
):
    recipients = [r.strip() for r in payload.recipients if r.strip()]
    cc_list = [c.strip() for c in payload.cc_list if c.strip()]

    # Always enforce mandatory CC
    for mandatory_cc in GLOBAL_MANDATORY_CC:
        if mandatory_cc not in cc_list:
            cc_list.append(mandatory_cc)

    if not recipients and not cc_list:
        raise HTTPException(
            status_code=400,
            detail="At least one recipient or CC email must be provided.",
        )

    try:
        # 1. Build Excel file and parse dataset
        excel_bytes, filename, headers, data_rows, summary_cards = (
            generate_excel_bytes(payload, user)
        )

        # 2. Format Dates in Subject (e.g. 23 July 2026 to 29 July 2026)
        start_fmt = format_display_date(payload.start_date)
        end_fmt = format_display_date(payload.end_date)

        if start_fmt and end_fmt:
            date_range_str = f"{start_fmt} to {end_fmt}"
        elif start_fmt:
            date_range_str = f"From {start_fmt}"
        elif end_fmt:
            date_range_str = f"Up to {end_fmt}"
        else:
            date_range_str = "All Time"

        report_title = payload.report_type.replace("_", " ").title()
        subject = f"Teleglobal NOC Report: {report_title} ({date_range_str})"

        # 3. Render HTML Template
        template = email_jinja_env.get_template("report_email.html")
        html_body = template.render(
            report_type_title=report_title,
            date_range_str=date_range_str,
            team_filter=payload.team or "All Teams",
            operator_name=user.get("full_name", "NOC Operator"),
            summary_cards=summary_cards,
            headers=headers,
            data_rows=data_rows,
        )

        # 4. Offload transmission to Background Task
        background_tasks.add_task(
            send_email_in_background,
            recipients,
            cc_list,
            subject,
            html_body,
            excel_bytes,
            filename,
        )

        return {
            "status": "success",
            "message": "Report dispatch request accepted! Sending email in background.",
        }

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to prepare report mail: {str(e)}"
        )

# =========================================================================
# SHIFT HANDOVER / TICKET FORWARDING ENDPOINTS
# =========================================================================

class TicketForwardPayload(BaseModel):
    ticket_id: int
    forward_to_username: str
    remarks: Optional[str] = ""

@app.get("/api/users/list")
async def get_users_list(current_user=Depends(get_current_user)):
    """
    Returns active user directory (id, username, full_name, email_id).
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT 
                id, 
                username, 
                COALESCE(full_name, username) AS full_name, 
                email_id AS email 
            FROM users 
            ORDER BY full_name ASC
        """)
        users = cursor.fetchall()
        return users
    except Exception as e:
        print(f"[ERROR] /api/users/list: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch user directory: {str(e)}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.post("/api/tickets/forward")
async def forward_ticket(
    payload: TicketForwardPayload, 
    background_tasks: BackgroundTasks,
    current_user=Depends(get_current_user)
):
    """
    Forwards a ticket to a new shift engineer by updating assigned_to_name
    while permanently leaving open_by_name unchanged.
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Sanitize ticket_id into an integer
        raw_val = str(payload.ticket_id).strip()
        numeric_digits = ''.join(filter(str.isdigit, raw_val))
        clean_ticket_id = int(numeric_digits) if numeric_digits else None

        if clean_ticket_id is None:
            raise HTTPException(status_code=400, detail=f"Invalid Ticket ID format: {payload.ticket_id}")

        # 1. Fetch current ticket details strictly using 'ticket_id'
        cursor.execute("SELECT * FROM tickets WHERE ticket_id = %s", (clean_ticket_id,))
        ticket = cursor.fetchone()

        if not ticket:
            raise HTTPException(status_code=404, detail=f"Ticket #{clean_ticket_id} not found in database.")

        # 2. Fetch target user details from users table using username or full_name
        cursor.execute(
            "SELECT full_name, username, email_id FROM users WHERE username = %s OR full_name = %s", 
            (payload.forward_to_username, payload.forward_to_username)
        )
        target_user = cursor.fetchone()

        new_assignee_name = target_user['full_name'] if (target_user and target_user.get('full_name')) else payload.forward_to_username
        forwarder_name = current_user.get('full_name') or current_user.get('username') or "NOC Engineer"

        # Format Handover History Remarks
        formatted_remark = (
            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
            f"Handed over from '{forwarder_name}' to '{new_assignee_name}': {payload.remarks}"
        )

        existing_remarks = ticket.get('forwarding_remarks') or ""
        updated_remarks = f"{existing_remarks}\n{formatted_remark}".strip()

        # 3. Update ONLY assigned_to_name & forwarding_remarks (DO NOT touch open_by_name)
        cursor.execute(
            """
            UPDATE tickets 
            SET assigned_to_name = %s,
                forwarding_remarks = %s
            WHERE ticket_id = %s
            """,
            (new_assignee_name, updated_remarks, clean_ticket_id)
        )
        conn.commit()

        # 4. Email Notification
        target_email = target_user.get('email_id') if target_user else None
        if target_email:
            subject = f"[SHIFT HANDOVER] Ticket #{clean_ticket_id} Forwarded to You | Circuit: {ticket.get('circuit_id', 'N/A')}"
            body = f"""
            <h3>Shift Handover Notification</h3>
            <p>Ticket <b>#{clean_ticket_id}</b> has been forwarded to you by <b>{forwarder_name}</b>.</p>
            <p><b>Original Creator:</b> {ticket.get('open_by_name', 'N/A')}</p>
            <p><b>Circuit ID:</b> {ticket.get('circuit_id', 'N/A')}</p>
            <p><b>Issue Category:</b> {ticket.get('issue_category', 'N/A')}</p>
            <p><b>Handover Remarks / Progress:</b></p>
            <blockquote style="background:#f1f5f9; padding:10px; border-left:4px solid #2563eb;">
                {payload.remarks or 'No specific remarks provided.'}
            </blockquote>
            <p>Please review and continue tracking this incident.</p>
            """
            msg = build_email_message(SMTP_USER, target_email, GLOBAL_MANDATORY_CC, subject, body)
            background_tasks.add_task(send_smtp_email_background, msg.as_string(), [target_email] + GLOBAL_MANDATORY_CC)

        return {"status": "success", "message": f"Ticket #{clean_ticket_id} successfully assigned to {new_assignee_name}"}

    except HTTPException as he:
        raise he
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"[ERROR] /api/tickets/forward: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to forward ticket: {str(e)}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()